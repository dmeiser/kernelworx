"""Smoke tests for the shared-campaign reports page.

Covers issue #84: create and join a shared campaign to generate data, then
navigate to ``/campaign-reports`` and verify Unit Summary, Seller Report, and
Order Details render with rollup values tied to the seeded order and that both
Excel export buttons produce downloads.
"""

import re
import time
import urllib.parse
import uuid
from pathlib import Path

import pytest
from openpyxl import load_workbook
from playwright.sync_api import Browser, BrowserContext, Page, expect

from tests.e2e.pages.campaign_reports_page import CampaignReportsPage
from tests.e2e.pages.order_page import OrderPage
from tests.e2e.pages.shared_campaigns_page import SharedCampaignsPage
from tests.e2e.utils.auth import login_as_owner

# ---------------------------------------------------------------------------
# Private helpers
# ---------------------------------------------------------------------------


def _unique_campaign_name(base: str) -> str:
    """Return a campaign name whose first four characters are unique.

    The shared-campaign code uses the first four characters of the campaign
    name as an abbreviation, so reuse of the same prefix causes a code
    collision in the dev environment.
    """
    prefix = f"E{uuid.uuid4().hex[:3].upper()}"
    return f"{prefix} {base} {int(time.time())}"


def _code_for_campaign_name(page: Page, campaign_name: str) -> str:
    """Return the short code for the visible row matching *campaign_name*."""
    row = page.get_by_role("row").filter(has_text=campaign_name)
    expect(row).to_be_visible(timeout=10_000)
    code = row.get_by_role("cell").filter(has_text=re.compile(r"^[A-Z0-9]+(-[A-Z0-9]+)+$")).inner_text()
    return code


def _get_profile_id_from_url(url: str) -> str:
    """Extract the profile ID from a ``/scouts/{id}/…`` URL."""
    match = re.search(r"/scouts/([^/]+)", url)
    assert match, f"Could not extract profile_id from URL: {url}"
    return urllib.parse.unquote(match.group(1))


def _get_campaign_id_from_url(url: str) -> str:
    """Extract the campaign ID from a ``/scouts/{id}/campaigns/{id}/…`` URL."""
    match = re.search(r"/scouts/[^/]+/campaigns/([^/?#]+)", url)
    assert match, f"Could not extract campaign_id from URL: {url}"
    return urllib.parse.unquote(match.group(1))


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------


@pytest.mark.smoke
@pytest.mark.slow
def test_campaign_reports_generate_and_views(
    owner_page: Page, browser: Browser, ensure_owner_profile: str, tmp_path: Path
) -> None:
    """Create a shared campaign, join it, add an order, and verify reports."""
    # Step 1 — create a shared campaign as the owner.
    shared = SharedCampaignsPage(owner_page)
    shared.goto()
    shared.click_create()
    campaign_name = _unique_campaign_name("Unit Report")
    shared.create_shared_campaign(campaign_name=campaign_name)
    code = _code_for_campaign_name(owner_page, campaign_name)

    # Step 2 — join the shared campaign in a fresh context and capture IDs.
    join_context: BrowserContext = browser.new_context(ignore_https_errors=True)
    join_page: Page = join_context.new_page()
    seeded_order_total = ""
    customer_name = ""
    try:
        login_as_owner(join_page)
        join_shared = SharedCampaignsPage(join_page)
        join_shared.join_shared_campaign(code)
        assert "/campaigns/" in join_page.url, f"Expected campaign detail URL after joining; got: {join_page.url}"
        profile_id = _get_profile_id_from_url(join_page.url)
        campaign_id = _get_campaign_id_from_url(join_page.url)

        # Step 3 — create an order so the unit report has seller data.
        order_page = OrderPage(join_page)
        order_page.goto(profile_id, campaign_id)
        customer_name = f"Unit Report {int(time.time())}"
        order_page.create_order_first_product(customer_name, qty=1)
        seeded_order_total = order_page.get_list_total_for_customer(customer_name)
    finally:
        join_context.close()

    # Step 4 — generate the campaign report.
    reports = CampaignReportsPage(owner_page)
    reports.goto()
    assert reports.is_visible(), "Shared Campaign Reports heading must be visible"
    reports.select_campaign_by_code(code)
    assert reports.can_generate_report(), "Generate Report button must be enabled after selection"
    reports.generate_report()

    # Step 5 — assert Unit Summary renders with rollup values tied to the seeded order.
    assert reports.report_header_is_visible(), "Report header card must be visible after generation"
    assert reports.unit_summary_is_visible(), "Unit Summary cards and Top Sellers table must render"

    total_sellers = reports.get_rollup_value("Total Sellers")
    total_orders = reports.get_rollup_value("Total Orders")
    total_sales = reports.get_rollup_value("Total Sales")
    assert total_sellers and int(total_sellers) >= 1, f"Total Sellers must be at least 1; got: {total_sellers!r}"
    assert total_orders and int(total_orders) >= 1, f"Total Orders must be at least 1; got: {total_orders!r}"
    assert total_sales and total_sales.startswith("$"), (
        f"Total Sales must be a non-empty dollar value; got: {total_sales!r}"
    )
    if seeded_order_total:
        assert total_sales == seeded_order_total, (
            f"Total Sales {total_sales!r} must match seeded order total {seeded_order_total!r}"
        )
    assert reports.get_top_sellers_row_count() >= 1, "Top Sellers table must contain at least one row"

    # Step 6 — assert Seller Report table renders and has rows/data.
    reports.switch_to_seller_report()
    assert reports.seller_report_is_visible(), "Seller Report section and table must render"
    assert reports.get_active_table_row_count() >= 1, "Seller Report table must contain at least one row"

    # Step 7 — assert Order Details table renders and contains the seeded customer.
    reports.switch_to_order_details()
    assert reports.order_details_is_visible(), "Order Details section and table must render"
    assert reports.get_active_table_row_count() >= 1, "Order Details table must contain at least one row"
    customer_cells = reports.get_active_table_cell_texts("Customer")
    assert customer_name in customer_cells, (
        f"Order Details must include the seeded customer {customer_name!r}; got: {customer_cells}"
    )

    # Step 8 — invoke the export helpers and verify both downloads are non-empty XLSX files.
    seller_report_path = reports.download_seller_report_to(tmp_path / "seller_report.xlsx")
    assert seller_report_path.suffix == ".xlsx", f"Expected .xlsx seller report; got: {seller_report_path}"
    assert seller_report_path.stat().st_size > 0, "Downloaded Seller Report must not be empty"
    seller_workbook = load_workbook(seller_report_path)
    assert seller_workbook.active is not None, "Seller Report workbook must have an active worksheet"
    assert seller_workbook.active.max_row >= 2, "Seller Report must contain header and at least one data row"

    order_details_path = reports.download_order_details_to(tmp_path / "order_details.xlsx")
    assert order_details_path.suffix == ".xlsx", f"Expected .xlsx order details; got: {order_details_path}"
    assert order_details_path.stat().st_size > 0, "Downloaded Order Details must not be empty"
    order_workbook = load_workbook(order_details_path)
    assert order_workbook.active is not None, "Order Details workbook must have an active worksheet"
    assert order_workbook.active.max_row >= 2, "Order Details must contain header and at least one data row"
