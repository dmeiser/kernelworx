"""Smoke tests for campaign report generation and download."""

import csv
import re
import urllib.parse
from datetime import datetime
from pathlib import Path

import pytest
from openpyxl import load_workbook
from playwright.sync_api import Page

from tests.e2e.pages.campaign_page import CampaignPage
from tests.e2e.pages.dashboard_page import DashboardPage
from tests.e2e.pages.order_page import OrderPage
from tests.e2e.pages.reports_page import ReportsPage

# ---------------------------------------------------------------------------
# Private helpers
# ---------------------------------------------------------------------------


def _get_profile_id_from_url(url: str) -> str:
    """Extract the profile ID from a ``/scouts/{id}/…`` URL.

    Args:
        url: Full browser URL string.

    Returns:
        URL-decoded profile identifier.

    Raises:
        AssertionError: When the URL does not contain the expected pattern.
    """
    match = re.search(r"/scouts/([^/]+)", url)
    assert match, f"Could not extract profile_id from URL: {url}"
    return urllib.parse.unquote(match.group(1))


def _get_campaign_id_from_url(url: str) -> str:
    """Extract the campaign ID from a ``/scouts/{id}/campaigns/{id}/…`` URL.

    Args:
        url: Full browser URL string.

    Returns:
        URL-decoded campaign identifier.

    Raises:
        AssertionError: When the URL does not contain the expected pattern.
    """
    match = re.search(r"/scouts/[^/]+/campaigns/([^/?#]+)", url)
    assert match, f"Could not extract campaign_id from URL: {url}"
    return urllib.parse.unquote(match.group(1))


def _ensure_campaign_has_orders(page: Page, profile_id: str, campaign_id: str) -> None:
    """Create a sample order when the campaign has none.

    The Reports page only renders CSV/XLSX download buttons when the campaign
    has at least one order.
    """
    order_page = OrderPage(page)
    order_page.goto(profile_id, campaign_id)
    no_orders_message = page.get_by_text("No orders yet")
    if no_orders_message.is_visible():
        customer = f"Report Smoke {datetime.now().strftime('%Y%m%d%H%M%S')}"
        order_page.create_order_first_product(customer, qty=1)


def _ensure_owner_has_campaign(page: Page) -> tuple[str, str]:
    """Ensure the owner has a campaign and return its profile/campaign IDs."""
    dashboard = DashboardPage(page)
    dashboard.goto()
    dashboard.wait_for_profiles_loaded()
    profile_names = dashboard.get_profile_names()
    assert profile_names, "Owner must have at least one seller profile"
    dashboard.click_profile(profile_names[0])
    page.wait_for_url("**/campaigns**", timeout=10_000)
    profile_id = _get_profile_id_from_url(page.url)
    campaign_page = CampaignPage(page)

    campaign_names = campaign_page.get_campaign_names()
    if not campaign_names:
        campaign_name = f"Reports Download Test {datetime.now().strftime('%Y%m%d%H%M%S')}"
        campaign_page.create_campaign_first_catalog(campaign_name)
        campaign_names = campaign_page.get_campaign_names()
    assert campaign_names, "Need at least one campaign to test reports download"

    campaign_page.click_campaign(campaign_names[0])
    campaign_id = _get_campaign_id_from_url(page.url)
    return profile_id, campaign_id


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------


@pytest.mark.smoke
def test_campaign_reports_download_buttons(owner_page: Page, ensure_owner_profile: str) -> None:
    """Verify the Reports tab shows the order table and CSV/XLSX download buttons.

    Navigation strategy:

    1. Open the owner's dashboard and click the first profile.
    2. Use the first existing campaign, or create one if the list is empty.
    3. Click *View Orders* to land on the campaign detail page and capture the
       campaign ID from the URL.
    4. Navigate to the ``/reports`` tab.
    5. If the campaign has no orders, the download buttons do not render; in
       that case create a sample order and return to the reports tab.
    6. Assert the heading, *All Orders* table heading, and both download
       buttons are visible and enabled.
    7. Use Playwright's ``expect_download`` to confirm the CSV download starts.
    """
    # Step 1 – navigate to first profile's campaigns.
    dashboard = DashboardPage(owner_page)
    dashboard.goto()
    dashboard.wait_for_profiles_loaded()
    profile_names = dashboard.get_profile_names()
    assert profile_names, "Owner must have at least one seller profile"
    dashboard.click_profile(profile_names[0])
    owner_page.wait_for_url("**/campaigns**", timeout=10_000)
    profile_id = _get_profile_id_from_url(owner_page.url)
    campaign_page = CampaignPage(owner_page)

    # Step 2 – ensure at least one campaign exists.
    campaign_names = campaign_page.get_campaign_names()
    if not campaign_names:
        campaign_name = f"Reports Download Test {datetime.now().strftime('%Y%m%d%H%M%S')}"
        campaign_page.create_campaign_first_catalog(campaign_name)
        campaign_names = campaign_page.get_campaign_names()
    assert campaign_names, "Need at least one campaign to test reports download"

    # Step 3 – click into the campaign to capture its ID.
    campaign_page.click_campaign(campaign_names[0])
    campaign_id = _get_campaign_id_from_url(owner_page.url)

    # Step 4 – navigate to the reports tab.
    reports = ReportsPage(owner_page)
    reports.goto(profile_id, campaign_id)

    # Step 5 – ensure the campaign has orders so download buttons render.
    if not reports.download_csv_button_is_visible():
        _ensure_campaign_has_orders(owner_page, profile_id, campaign_id)
        reports.goto(profile_id, campaign_id)

    # Step 6 – assert the reports UI is visible and buttons are clickable.
    assert "/reports" in owner_page.url, f"Expected /reports in URL; got: {owner_page.url}"
    assert reports.heading_is_visible(), "Expected h5 containing 'Reports' to be visible on the reports tab"
    assert reports.table_heading_is_visible(), "Expected 'All Orders' heading to be visible on the reports tab"
    assert reports.download_csv_button_is_visible(), "Expected CSV download button to be visible"
    assert reports.download_xlsx_button_is_visible(), "Expected XLSX download button to be visible"
    assert reports.download_csv_button_is_enabled(), "Expected CSV download button to be enabled"
    assert reports.download_xlsx_button_is_enabled(), "Expected XLSX download button to be enabled"

    # Step 7 – verify a CSV download starts.
    with owner_page.expect_download() as download_info:
        reports.click_download_csv()
    download = download_info.value
    assert download.suggested_filename.endswith(".csv"), f"Expected CSV download; got: {download.suggested_filename}"


@pytest.mark.smoke
def test_download_xlsx(owner_page: Page, ensure_owner_profile: str, tmp_path: Path) -> None:
    """Create an order, download XLSX, and verify its contents with openpyxl."""
    profile_id, campaign_id = _ensure_owner_has_campaign(owner_page)

    order_page = OrderPage(owner_page)
    order_page.goto(profile_id, campaign_id)
    customer_name = f"XLSX Smoke {datetime.now().strftime('%Y%m%d%H%M%S')}"
    order_page.create_order_first_product(customer_name, qty=1)

    reports = ReportsPage(owner_page)
    reports.goto(profile_id, campaign_id)
    assert reports.download_xlsx_button_is_visible(), "XLSX download button must be visible"
    assert reports.download_xlsx_button_is_enabled(), "XLSX download button must be enabled"

    xlsx_path = reports.download_xlsx_to(tmp_path / "report.xlsx")
    assert xlsx_path.suffix == ".xlsx", f"Expected .xlsx file; got: {xlsx_path}"
    assert xlsx_path.stat().st_size > 0, "Downloaded XLSX file must not be empty"

    workbook = load_workbook(xlsx_path)
    worksheet = workbook.active
    assert worksheet is not None, "XLSX workbook must have an active worksheet"
    headers = [worksheet.cell(row=1, column=col).value for col in range(1, worksheet.max_column + 1)]
    assert headers[:3] == ["Name", "Phone", "Address"], f"Unexpected CSV headers: {headers}"
    assert headers[-1] == "Total", f"Expected 'Total' as last header; got: {headers[-1]}"

    customer_cells = [
        worksheet.cell(row=row, column=1).value for row in range(2, worksheet.max_row + 1)
    ]
    assert customer_name in customer_cells, "XLSX must contain the smoke-test order row"


@pytest.mark.smoke
def test_csv_content_parsed(owner_page: Page, ensure_owner_profile: str, tmp_path: Path) -> None:
    """Download a CSV report and parse it to verify the smoke-test order row."""
    profile_id, campaign_id = _ensure_owner_has_campaign(owner_page)

    order_page = OrderPage(owner_page)
    order_page.goto(profile_id, campaign_id)
    customer_name = f"CSV Smoke {datetime.now().strftime('%Y%m%d%H%M%S')}"
    order_page.create_order_first_product(customer_name, qty=1)

    reports = ReportsPage(owner_page)
    reports.goto(profile_id, campaign_id)
    assert reports.download_csv_button_is_visible(), "CSV download button must be visible"
    assert reports.download_csv_button_is_enabled(), "CSV download button must be enabled"

    csv_path = reports.download_csv_to(tmp_path / "report.csv")
    with csv_path.open(newline="") as f:
        rows = list(csv.reader(f))

    assert rows, "CSV file must contain at least a header row"
    headers = rows[0]
    assert headers[:3] == ["Name", "Phone", "Address"], f"Unexpected CSV headers: {headers}"
    assert headers[-1] == "Total", f"Expected 'Total' as last header; got: {headers[-1]}"

    matching_rows = [row for row in rows[1:] if row and row[0] == customer_name]
    assert matching_rows, "CSV must contain the smoke-test order row"
    order_total = matching_rows[0][-1]
    assert order_total, "Smoke-test order row must have a Total value"
